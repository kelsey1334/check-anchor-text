import os
import asyncio
import aiohttp
import logging
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
from telegram import Update, InputFile
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
)

load_dotenv()
BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")

logging.basicConfig(level=logging.INFO)

USER_TASKS = {}

def detect_page_type(soup):
    if soup.find('div', class_='article-inner'):
        return 'post'
    elif soup.find('div', id='content'):
        return 'page'
    elif soup.find('div', class_='taxonomy-description'):
        return 'category'
    return None

def extract_content(soup, page_type):
    if page_type == 'post':
        node = soup.find('div', class_='article-inner')
    elif page_type == 'page':
        node = soup.find('div', id='content')
    elif page_type == 'category':
        node = soup.find('div', class_='taxonomy-description')
    else:
        return ""
    return node.decode_contents() if node else ""

def find_internal_links(html, base_url):
    soup = BeautifulSoup(html, "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        anchor = a.text.strip()
        if not anchor:
            continue
        if (href.startswith("/") or base_url in href):
            full_link = href if href.startswith("http") else base_url.rstrip("/") + "/" + href.lstrip("/")
            links.append((anchor, full_link))
    return links

async def check_link_status(session, url):
    try:
        async with session.get(url, allow_redirects=True, timeout=10) as response:
            return response.status
    except Exception:
        return None

async def process_url(session, url, stt, page_type=None):
    try:
        async with session.get(url, timeout=15) as resp:
            if resp.status != 200:
                return [stt, url, "Toàn bộ URL lỗi", url, "", "", ""]
            html = await resp.text()
    except Exception:
        return [stt, url, "Không truy cập được URL", url, "", "", ""]
    
    soup = BeautifulSoup(html, "html.parser")
    # Nếu page_type không có, tự động detect
    if not page_type:
        page_type = detect_page_type(soup)
    if not page_type:
        return [stt, url, "Không nhận diện dạng bài", url, "", "", ""]
    content = extract_content(soup, page_type)
    if not content:
        return [stt, url, "Không lấy được nội dung", url, "", "", ""]
    base_url = url.split("//", 1)[-1].split("/", 1)[0]
    base_url = url.split("//")[0] + "//" + base_url

    internal_links = find_internal_links(content, base_url)
    result_row = [stt, url]
    anchor_error = 0
    for anchor, link in internal_links:
        code = await check_link_status(session, link)
        if code in [301, 404]:
            anchor_error += 1
            result_row.append(f"{anchor} - {link}")
            if anchor_error == 4:
                break
    while len(result_row) < 6:
        result_row.append("")
    return result_row

async def handle_excel(
    file_path, output_path, context, chat_id, user_id
):
    wb = load_workbook(file_path)
    ws = wb.active
    # Lấy header (dòng đầu tiên)
    headers = [cell.value for cell in ws[1]]
    url_idx = None
    type_idx = None
    for idx, val in enumerate(headers):
        if val and str(val).strip().lower() == 'url':
            url_idx = idx
        if val and str(val).strip().lower() == 'type':
            type_idx = idx
    if url_idx is None:
        await context.bot.send_message(chat_id=chat_id, text="Không tìm thấy cột 'url' trong file.")
        return

    urls_types = []
    for row in ws.iter_rows(min_row=2, max_col=max(url_idx, type_idx if type_idx is not None else 0)+1):
        url = row[url_idx].value if url_idx < len(row) else None
        page_type = None
        if type_idx is not None and type_idx < len(row):
            page_type = row[type_idx].value
            if page_type:
                page_type = str(page_type).strip().lower()
        if url:
            urls_types.append((url, page_type))

    result_wb = Workbook()
    result_ws = result_wb.active
    result_ws.append([
        'stt', 'url check', 
        'anchor bị lỗi và link lỗi 1', 'anchor bị lỗi và link lỗi 2', 
        'anchor bị lỗi và link lỗi 3', 'anchor bị lỗi và link lỗi 4'
    ])

    async with aiohttp.ClientSession() as session:
        for i, (url, page_type) in enumerate(urls_types, 1):
            if USER_TASKS.get(user_id, {}).get("cancel", False):
                await context.bot.send_message(chat_id=chat_id, text=f"Đã dừng tiến trình theo yêu cầu! Dừng ở dòng {i}/{len(urls_types)}.")
                break
            await context.bot.send_message(chat_id=chat_id, text=f"Đang kiểm tra {i}/{len(urls_types)}: {url} ({page_type if page_type else 'auto'})")
            row = await process_url(session, url, i, page_type)
            result_ws.append(row)
    result_wb.save(output_path)
    # In debug để xác nhận file thực sự có data trước khi gửi về
    print(f"Saved output: {output_path}")
    print(f"File size: {os.path.getsize(output_path)} bytes")

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Gửi file Excel (.xlsx) chứa danh sách URL cần kiểm tra.\n"
        "File nên gồm 2 cột: url, type (post/page/category).\n"
        "Nếu không có cột 'type', bot sẽ tự đoán loại trang.\n"
        "Gửi /cancel để dừng tiến trình."
    )

async def handle_doc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    chat_id = update.effective_chat.id

    USER_TASKS[user_id] = {"cancel": False}

    if update.message.document.mime_type not in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
        await update.message.reply_text("Chỉ nhận file .xlsx.")
        return
    file = await context.bot.get_file(update.message.document.file_id)
    input_path = f"input_{update.message.document.file_id}.xlsx"
    output_path = f"output_{update.message.document.file_id}.xlsx"
    await file.download_to_drive(input_path)

    await update.message.reply_text("Đang xử lý, vui lòng đợi... (có thể gửi /cancel để dừng lại)")

    try:
        await handle_excel(input_path, output_path, context, chat_id, user_id)
        # Kiểm tra file output thực tế có tồn tại và có dữ liệu không
        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
            await update.message.reply_document(InputFile(output_path, filename="ketqua_internal_link.xlsx"))
        else:
            await context.bot.send_message(chat_id=chat_id, text="Không tạo được file kết quả hoặc file kết quả rỗng.")
    except Exception as ex:
        await context.bot.send_message(chat_id=chat_id, text=f"Lỗi: {ex}")
    finally:
        if os.path.exists(input_path): os.remove(input_path)
        if os.path.exists(output_path): os.remove(output_path)
        USER_TASKS.pop(user_id, None)  # cleanup

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id in USER_TASKS:
        USER_TASKS[user_id]["cancel"] = True
        await update.message.reply_text("Đã nhận lệnh dừng tiến trình.")
    else:
        await update.message.reply_text("Không có tiến trình nào đang chạy.")

def main():
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("cancel", cancel))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_doc))
    print("Bot is running...")
    app.run_polling()

if __name__ == "__main__":
    main()
